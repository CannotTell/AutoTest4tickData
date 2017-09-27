#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 8/31/2017 5:25 PM
# @Author  : JZK
# @File    : Check_Data.py

from openpyxl import load_workbook
import datetime
import math
import os



class tick_data_check:
    def __init__(self):
        pass
    def get_Data(self,sourceFile):
        wb = load_workbook(filename = sourceFile,read_only = True)
        ws = wb.active
        rows = ws.rows
        content = []
        for row in rows:
            ll = []
            for k, cell in enumerate(row):
                ll.append(cell.value)
                if (k + 1) % ws.max_column == 0:
                    content.append(ll)
                    ll = []
        return content


    def getDataCompInfo(self, SourceDataList, genDataList, fileName):
        count = 0

        for i in genDataList:
            if (type(i[1]) is datetime.datetime):
                date = i[1].strftime("%Y/%m/%d")
                time = i[2].strftime("%H:%M")
                if i[3] is not None:
                    price = i[3]
                elif i[4] is not None:
                    price = i[4]
                else:
                    price = i[5]
                # print(date,time, price)
                for k, v in enumerate(SourceDataList):
                    if v[0] is not None:
                        strs = v[0].strip().split('-')

                        if date == strs[0] and time == strs[1]:
                            count += 1
                            # print(abs(v[4] - price) / price)
                            # print(v[4], price)
                            if abs(v[4] - price) / price > 0.05:
                            # if abs(v[4] - price)  > 0.05:
                                info = '{}:{}-{}:SourceData:{}, GenData:{}'.format(fileName, date, time, v[4], price)
                                self.RecordData(info)
                    else:
                        del SourceDataList[k]
                        continue
        return count

    def RecordData(self, info):
        with open('info.txt','a') as f:
            f.write(info)
            f.write('\n')



if __name__ == '__main__':
    folder_path = 'C:\\Users\\zhukai.jiang\\Desktop\\DownloadStock'
    os.chdir(folder_path)
    fileName_list = os.listdir()
    tdc = tick_data_check()

    source_path = 'C:\\Users\\zhukai.jiang\\Desktop\\600848.xlsx'

    for i in fileName_list:
        genFile_path = os.path.join(folder_path, i)


        source_data = tdc.get_Data(source_path)
        gen_data = tdc.get_Data(genFile_path)
        # for i in gen_data:
        #     print(i)
        count = tdc.getDataCompInfo(source_data, gen_data, i)

        print('{}文件已匹配完{}条数据'.format(i,count))

